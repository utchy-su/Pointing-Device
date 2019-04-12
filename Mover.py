# -*- coding: utf-8 -*-
"""
Created on Mon Jun 11 14:06:55 2018

@author: smart
"""

# coding:utf-8
import serial
import win32api
import math
import openpyxl as px
import keyboard
import numpy as np
# from HPF import *


class Function:
    def __init__(self, user_name='PERSON', mode='linear'):
        self.ser = serial.Serial('COM15', 19200)
        self.gyro_pitch = 0
        self.gyro_roll = 0
        self.gyro_yaw = 0
        self.dt = 0.0316
        self.user_name = user_name
        self.mode = mode
        # self.filter = HPF()
        # open the serial port

        users_preset = {'PERSON': {'right_limit': 40, 'left_limit': 40, 'sigmoid_const': 0.25, 'forward_limit': 40,
                                   'backward_limit': 40}
                        }

        for key, value in users_preset.items():
            if key == user_name:
                self.right_limit = value['right_limit']
                self.left_limit = value['left_limit']
                self.a = value['sigmoid_const']
                self.forward_limit = value['forward_limit']
                self.backward_limit = value['backward_limit']

    def __getter_angle(self):
        # privatized method
        # make this public method if necessary
        data = self.ser.readline()
        data = data.decode().split(';')
        print(data)

        try:
            # ax = (6.091 * 10**-5) * (int(data[0]) - 1282.3)
            # ay = (6.059 * 10**-5) * (int(data[1]) - 92.60)
            # az = (6.050 * 10**-5) * (int(data[2]) + 535.5)
            # t = float(data[6])
            # print(ax, ay, az)
            ax = float(data[0])
            ay = float(data[1])
            az = float(data[2])
            t = float(data[3])
            print(ax, ay, az)

        except:
            ax, ay, az, t = 0, 0, 0, 0

        acc_pitch = math.atan2(ax, math.sqrt(az**2 + ay**2)) * 360 / 2 / math.pi
        acc_roll = math.atan2(ay, math.sqrt(ax**2+az**2)) * 360 / 2 / math.pi
        # angle calculated by accleration

        return acc_roll*-1, acc_pitch*-1, t

    def __mover_linear(self, x_input, y_input, kx, ky):
        x_now, y_now = win32api.GetCursorPos()

        if -self.left_limit <= x_input < 0:
            x_move = kx * x_input/self.left_limit
        elif 0 <= x_input < self.right_limit:
            x_move = kx * x_input / self.right_limit
        elif x_input < -self.left_limit:
            x_move = -kx
        elif self.right_limit < x_input:
            x_move = kx

        if -self.backward_limit <= y_input < 0:
            y_move = ky * y_input/self.backward_limit
        elif 0 <= y_input < self.forward_limit:
            y_move = ky * y_input/self.forward_limit
        elif y_input < -self.backward_limit:
            y_move = -ky
        elif self.right_limit < y_input:
            y_move = ky

        win32api.SetCursorPos((int(x_now + x_move), int(y_now + y_move)))

        # the command to move the cursor

        return int(x_now + x_move), int(y_now + y_move), x_now, y_now

    def __mover_sigmoid(self, x_input, y_input, kx, ky):
        x_now, y_now = win32api.GetCursorPos()

        if -self.left_limit <= x_input < -1:
            x_move = kx * (1/(1+math.e**(-self.a * (x_input + self.left_limit/2)))-1)
        elif -1 <= x_input < 1:
            x_move = 0
        elif 1 <= x_input <= self.right_limit:
            x_move = kx * 1/(1+math.e**(-self.a * (x_input - self.right_limit/2)))
        else:
            x_move = 0

        if -self.backward_limit <= y_input < -1:
            y_move = ky * (1/(1+math.e**(-self.a * (y_input + self.backward_limit/2)))-1)
        elif -1 <= y_input < 1:
            y_move = 0
        elif 1 <= y_input < self.forward_limit:
            y_move = ky * 1/(1+math.e**(-self.a * (y_input - self.forward_limit/2)))
        else:
            y_move = 0

        win32api.SetCursorPos((int(x_now + x_move), int(y_now + y_move)))

        return int(x_now + x_move), int(y_now + y_move), x_now, y_now

    def __mover_tanh(self, x_input, y_input, kx, ky):
        x_now, y_now = win32api.GetCursorPos()

        if -self.left_limit <= x_input < 0:
            x_move = kx * np.tanh(x_input/self.left_limit)
        elif 0 <= x_input < self.right_limit:
            x_move = kx * np.tanh(x_input/self.right_limit)
        elif x_input < -self.left_limit:
            x_move = kx * np.tanh(-1)
        elif self.right_limit < x_input:
            x_move = kx * np.tanh(1)

        if -self.backward_limit <= y_input < 0:
            y_move = ky * np.tanh(y_input/self.backward_limit)
        elif 0 <= y_input < self.forward_limit:
            y_move = ky * np.tanh(y_input/self.forward_limit)
        elif y_input < -self.backward_limit:
            y_move = ky * np.tanh(-1)
        elif self.forward_limit < y_input:
            y_move = ky * np.tanh(1)

        win32api.SetCursorPos((int(x_now + x_move), int(y_now + y_move)))
        return int(x_now + x_move), int(y_now + y_move), x_now, y_now

    def __mover_discrete(self, x_input, y_input, kx, ky):
        x_now, y_now = win32api.GetCursorPos()

        if -self.left_limit <= x_input < -self.left_limit * 2/3:
            x_move = kx * -1.0
        elif -self.left_limit * (2/3) <= x_input < -self.left_limit * (1/3):
            x_move = kx * -0.6
        elif -self.left_limit * (1/3) <= x_input < -2:
            x_move = kx * -0.3
        elif -2 <= x_input < 2:
            x_move = 0
        elif 2 <= x_input < self.right_limit * (1/3):
            x_move = kx * 0.3
        elif self.right_limit * (1/3) <= x_input < self.right_limit * (2/3):
            x_move = kx * 0.6
        elif self.right_limit * (2/3) <= x_input < self.right_limit:
            x_move = kx * 1.0
        else:
            x_move = 0

        if -self.backward_limit <= y_input < -1 * self.backward_limit*(2/3):
            y_move = ky * -1.0
        elif -self.backward_limit*(2/3) <= y_input < -self.backward_limit*(1/3):
            y_move = ky * -0.6
        elif -self.backward_limit*(1/3) <= y_input < -2:
            y_move = ky * -0.3
        elif -2 <= y_input < 2:
            y_move = 0
        elif 2 <= y_input < self.forward_limit * (1/3):
            y_move = ky * 0.3
        elif self.forward_limit * (1/3) <= y_input < self.forward_limit * (2/3):
            y_move = ky * 0.6
        elif self.forward_limit * (2/3) <= y_input < self.forward_limit:
            y_move = ky * 1.0
        else:
            y_move = 0

        win32api.SetCursorPos((int(x_now + x_move), int(y_now + y_move)))
        return int(x_now + x_move), int(y_now + y_move), x_now, y_now

    def __ex_saver(self, wb, ws, x_inp, y_inp, x_inp_fil, y_inp_fil, time, x, y, i):
        ws.cell(row=1, column=1).value = 'x_ang_filed'
        ws.cell(row=1, column=2).value = 'y_ang_filed'
        ws.cell(row=1, column=3).value = 'x_ang'
        ws.cell(row=1, column=4).value = 'y_ang'
        ws.cell(row=1, column=5).value = 'time'
        ws.cell(row=1, column=6).value = 'x_cod'
        ws.cell(row=1, column=7).value = 'y_cod'

        ws.cell(row=i+1, column=1).value = x_inp_fil
        ws.cell(row=i+1, column=2).value = y_inp_fil
        ws.cell(row=i+1, column=3).value = x_inp
        ws.cell(row=i+1, column=4).value = y_inp
        ws.cell(row=i+1, column=5).value = time
        ws.cell(row=i+1, column=6).value = x
        ws.cell(row=i+1, column=7).value = y

        return wb, ws

    # def tgt_func(self,x,a,b,c,d):
    #    return a*x**3 + b*x**2 + c*x**1 + d

    def calibrator(self):
        pitch = 0
        roll = 0
        for i in range(200):
            p_plus, r_plus, time = self.__getter_angle()
            pitch += p_plus
            roll += r_plus
            print('caliburation in progress...')
        pitch = pitch/200
        roll = roll/200

        return pitch, roll

    def main(self, data_path, filter_switch):
        print('press q when quitting')
        wb = px.Workbook()
        ws = wb.active
        kx, ky = 15, 15
        i = 1
        init_1, init_2 = self.calibrator()
        print('#############################')
        print('User: ', self.user_name)
        print('Mode: ', self.mode)
        if filter_switch == 1:
            # High Pass Filterを昔実装していました．今はついてません．
            pass

        elif filter_switch == 0:
            if self.mode == 'linear':
                while True:
                    y_inp, x_inp, t = self.__getter_angle()
                    x, y, x_now, y_now = self.__mover_linear(x_inp, y_inp - init_1, kx, ky)
                    wb, ws = self.__ex_saver(wb, ws, x_inp, y_inp, 'N/A', 'N/A', t, x, y, i)
                    i += 1
                    if keyboard.is_pressed('esc') == 1:
                        wb.save(data_path)
                        print('THANKS')
                        quit(0)
                        break

            elif self.mode == 'sigmoid':
                while True:
                    y_inp, x_inp, t = self.__getter_angle()
                    x, y, x_now, y_now = self.__mover_sigmoid(x_inp, y_inp - init_1, kx, ky)
                    wb, ws = self.__ex_saver(wb, ws, x_inp, y_inp, 'N/A', 'N/A', t, x, y, i)
                    i += 1
                    if keyboard.is_pressed('esc') == 1:
                        wb.save(data_path)
                        print('THANKS')
                        quit(0)
                        break

            elif self.mode == 'tanh':
                while True:
                    y_inp, x_inp, t = self.__getter_angle()
                    x, y, x_now, y_now = self.__mover_tanh(x_inp, y_inp - init_1, kx, ky)
                    wb, ws = self.__ex_saver(wb, ws, x_inp, y_inp, 'N/A', 'N/A', t, x, y, i)
                    i += 1
                    if keyboard.is_pressed('esc') == 1:
                        wb.save(data_path)
                        print('THANKS')
                        quit(0)
                        break

            elif self.mode == 'discrete':
                while True:
                    y_inp, x_inp, t = self.__getter_angle()
                    x, y, x_now, y_now = self.__mover_discrete(x_inp, y_inp - init_1, kx, ky)
                    wb, ws = self.__ex_saver(wb, ws, x_inp, y_inp, 'N/A', 'N/A', t, x, y, i)
                    i += 1
                    if keyboard.is_pressed('esc') == 1:
                        wb.save(data_path)
                        print('THANKS')
                        quit(0)
                        break

            else:
                print('error: incorrect mode selected')
                quit(1)


if __name__ == '__main__':
    data_path = 'PATH'

    test = Function(mode='linear')
    test.main(data_path, 0)
