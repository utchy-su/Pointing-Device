# -*- coding: utf-8 -*-
"""
Created on Mon Jun 11 14:02:26 2018

@author: smart
"""
# hello


import openpyxl as px
import math
import numpy as np
import matplotlib.pyplot as plt

class Analyzer:

    def __init__(self, data_path, path):
        self.data_path = data_path
        print(self.data_path)
        self.path = path
        self.wb = px.load_workbook(self.data_path)
        self.ws = self.wb.active

    def __data_getter(self):
        mat = []
        i = 0
        for column in self.ws.columns:
            plots = []
            for cell in column:
                try:
                    plots.append(int(cell.value))
                except:
                    plots.append(cell.value)
            if i == 0:
                mat.append(plots)
            else:
                mat.append(plots[::20])

        return np.array(mat)

    def __TRE_counter(self, mat):
        TRE = []
        for i in range(1, 15):
            tgt_num = [var for var in mat[0] if isinstance(var, str)==False]
            x_dst = 450 + 200*np.cos(math.pi*tgt_num[i]/8)
            y_dst = 450 + 200*np.sin(math.pi*tgt_num[i]/8)
            # print(tgt_num)

            dist = []
            inner_entry = 0
            outer_evac = 0
            # print('len=', len([var for var in mat[3*i+1] if var is not None]))
            for j in range(1, len([var for var in mat[3*i+1] if var is not None])):
                x_dif = mat[3*i+1][j] - x_dst
                y_dif = mat[3*i+2][j] - y_dst
                dist.append(math.sqrt(x_dif**2+y_dif**2))
            for j in range(1, len(dist)):
                if dist[j] <= 30 and dist[j-1] >= 30:
                    # inner entry
                    inner_entry += 1
                elif dist[j] >= 30 and dist[j-1] <= 30:
                    # outer evacuation
                    outer_evac += 1
                else:
                    pass

            TRE.append((inner_entry+outer_evac)//2)

        return TRE

    def __Ideal_route(self,mat,current_circle):
        x_from = int(450 + 200*np.cos(math.pi*mat[0][current_circle]/8))
        y_from = int(450 + 200*np.sin(math.pi*mat[0][current_circle]/8))

        x_tgt = int(450 + 200*np.cos(math.pi*mat[0][current_circle+1]/8))
        y_tgt = int(450 + 200*np.sin(math.pi*mat[0][current_circle+1]/8))

        # explicit function expression
        if x_from != x_tgt:
            tilt = (y_tgt - y_from)/(x_tgt - x_from)
        else:
            tilt = 'inf'

        return tilt

    def __Func(self, mat, i, x):
        x1 = 450 + 200*np.cos(math.pi*(mat[0][i]/8))
        y1 = 450 + 200*np.sin(math.pi*(mat[0][i]/8))
        tilt = self.__Ideal_route(mat,i)


        if tilt != 'inf':
            y = tilt*(x-x1) + y1
            return y, -1*tilt, tilt*x1-y1
        elif tilt == 'inf':
            y = 'undefinable'
            return y, None, None
            ############

    def __Across_counter(self, mat):
        ###############
        across_count = []

        for i in range(1,15):
            tmp = 0
            for j in range(1,len([var for var in mat[3*i+1] if var is not None])-1):
                x = mat[3*i+1][j]
                x_nxt = mat[3*i+1][j+1]
                y = mat[3*i+2][j]
                y_nxt = mat[3*i+2][j+1]
                F_x,_,_ = self.__Func(mat,i,x)
                F_x_nxt,_,_ = self.__Func(mat,i,x_nxt)

                if F_x == 'undefinable' or F_x_nxt == 'undefinable':
                    x_bs = 450 + 200 * np.cos(math.pi*mat[0][i]/8)
                    prev = x_bs - x
                    nxt = x_bs-x_nxt
                    # function is x = 45
                    if np.sign(prev) != np.sign(nxt):
                        tmp+=1
                    else:
                        pass
                else:
                    prev = F_x - y
                    nxt = F_x_nxt - y_nxt
                    if (np.sign(prev) != np.sign(nxt)) and (abs(prev) + abs(nxt) > 5):
                        tmp += 1
                    else:
                        pass

            across_count.append(tmp)

        ##################

        return across_count

    def __Orthogonal_direction(self, mat):
        ODC = []
        for i in range(1,15):
            x = [var for var in mat[3 * i + 1] if var is not None and isinstance(var, str) is False]
            y = [var for var in mat[3 * i + 2] if var is not None and isinstance(var, str) is False]

            x_from = int(450 + 200 * np.cos(math.pi * mat[0][i] / 8))
            y_from = int(450 + 200 * np.sin(math.pi * mat[0][i] / 8))

            x_tgt = int(450 + 200 * np.cos(math.pi * mat[0][i + 1] / 8))
            y_tgt = int(450 + 200 * np.sin(math.pi * mat[0][i + 1] / 8))

            vec1 = [x_tgt - x_from, y_tgt - y_from]
            tmp = 0
            cos_theta = []
            vec = []
            for j in range(len(x) - 1):
                if (x[j+1] - x[j]) != 0 or (y[j+1] - y[j]) != 0:
                    vec.append([x[j + 1] - x[j], y[j + 1] - y[j]])
            print("vec:", vec)
            vec_ave = []
            for j in range(0, len(vec)-5, 5):
                vec_ave.append([vec[j][0]+vec[j+1][0]+vec[j+2][0]+vec[j+3][0]+vec[j+4][0], vec[j][1]+vec[j+1][1]
                                +vec[j+2][1]+vec[j+3][1]+vec[j+4][1]])
            print("vec average: ", vec_ave)
            for j in range(len(vec_ave)):
                try:
                    cos_theta.append((vec_ave[j][0] * vec1[0] + vec_ave[j][1] * vec1[1]) / (
                                math.sqrt(vec_ave[j][0] ** 2 + vec_ave[j][1] ** 2) * math.sqrt(vec1[0] ** 2 + vec1[1] ** 2)))
                except ZeroDivisionError:
                    pass
            print("cos theta: ", cos_theta)
            print("len costheta: ", len(cos_theta))
            for k in range(len(cos_theta) - 1):
                if np.sign(cos_theta[k]) != np.sign(cos_theta[k + 1]):
                    tmp += 1
            print("tmp: ", tmp)

            ODC.append(tmp)

        return ODC

    def __Distance_from_line(self, mat):
        MDC = []
        for i in range(1,15):
            tmp = 0
            x = [var for var in mat[3*i+1] if var is not None and isinstance(var, str) == False]
            y = [var for var in mat[3*i+2] if var is not None and isinstance(var, str) == False]

            a = 1
            y_, b, c = self.__Func(mat,  i, 0)

            if y_== 'undefinable':
                x_bs = 450 + 200 * np.cos(math.pi*mat[0][i]/8)
                dist = [abs(var-x_bs) for var in x]
            else:
                dist = [abs(a*varx+b*vary+c)/math.sqrt(a**2+b**2) for varx, vary in zip(x,y)]

            dist_inc = []
            for j in range(len(dist)-1):
                dist_inc.append(dist[j+1]-dist[j])
            dist_inc = [var for var in dist_inc if var != 0]
            dist_inc_ave = []
            for j in range(0, len(dist_inc)-5, 5):
                dist_inc_ave.append(dist_inc[j] + dist_inc[j+1] + dist_inc[j+2] + dist_inc[j+3] + dist_inc[j+4])

            for k in range(len(dist_inc_ave)-1):
                if (np.sign(dist_inc_ave[k+1]) == -1) and (np.sign(dist_inc_ave[k]) == 1):
                    tmp += 1
                elif (np.sign(dist_inc_ave[k+1]) == 1) and (np.sign(dist_inc_ave[k]) == -1):
                    tmp += 1
            MDC.append(tmp)

        return MDC

    def __Movement_var_err_off(self,mat):
        MV = []
        ME = []
        MO = []
        for i in range(1,15):
            x = [var for var in mat[3*i+1] if var is not None and isinstance(var,str)==False]
            y = [var for var in mat[3*i+2] if var is not None and isinstance(var,str)==False]
            a = 1
            y_, b , c = self.__Func(mat,i,0)

            if y_=='undefinable':
                x_bs = 450 + 200 * np.cos(math.pi*mat[0][i]/8)
                dist = [abs(x_bs-var) for var in x]
            else:
                dist = [abs(a*varx+b*vary+c)/math.sqrt(a**2+b**2) for varx,vary in zip(x,y)]


            #############################
            if y_!='undefinable':
                Sum = 0
                for j in range(len(dist)):
                    a=1
                    _, b,c = self.__Func(mat,i,x[j])
                    z = a*y[j] + b*x[j] + c
                    Sum += np.sign(z)*dist[j]
                Ave = Sum / (len(dist))
                MO.append(Ave)
                Sigma = 0

                for k in range(len(dist)):
                    Sigma += (dist[k] - Ave)**2

                mv = math.sqrt(Sigma/(len(dist)-1))
                MV.append(mv)
            elif y_=='undefinable':
                Sum = 0
                for j in range(len(dist)):
                    Sum += np.sign(x[j]-450)*dist[j]
                Ave = Sum/len(dist)
                MO.append(Ave)
                Sigma = 0

                for k in range(len(dist)):
                    Sigma += (dist[k]-Ave)**2
                mv = math.sqrt(Sigma/(len(dist)-1))
                MV.append(mv)
            ############################
            Sum = 0
            a = 1
            for l in range(len(dist)):
                Sum += dist[l]
            me = Sum / len(dist)

            ME.append(me)
            ############################

        return MV, ME, MO

    def __Throughput_cal(self,mat):
        Throughput = []
        for i in range(1,15):
            tmp = 0
            time = [var for var in mat[3*(i+1)] if var is not None and isinstance(var,str)==False]
            time_init = time[0]*1/1000
            time_end = time[len(time)-1]*1/1000
            x1 = 450 + 200*np.cos(math.pi*(mat[0][i]/8))
            y1 = 450 + 200*np.sin(math.pi*(mat[0][i]/8))
            x2 = 450 + 200*np.cos(math.pi*(mat[0][i+1]/8))
            y2 = 450 + 200*np.sin(math.pi*(mat[0][i+1]/8))

            dist = math.sqrt((x2-x1)**2+(y2-y1)**2)
            tmp = np.log2(dist/10+1)/(time_end-time_init)

            Throughput.append(tmp)
        return Throughput

    """
    def __Effective_Throughput_cal(self, mat):
        Throughput = []
        x_scatter = []
        for i in range(1, 15):
            x1 = 450 + 200 * np.cos(math.pi * (mat[0][i] / 8))
            y1 = 450 + 200 * np.sin(math.pi * (mat[0][i] / 8))
            x2 = 450 + 200 * np.cos(math.pi * (mat[0][i + 1] / 8))
            y2 = 450 + 200 * np.sin(math.pi * (mat[0][i + 1] / 8))
            x = [var for var in mat[3*i+1] if var is not None and isinstance(var, str) is False]
            y = [var for var in mat[3*i+2] if var is not None and isinstance(var, str) is False]
            x_endclick = x[len(x) - 1]
            y_endclick = y[len(y) - 1]
            l = np.sqrt((x_endclick-x1)**2+(y_endclick-y1)**2)
            inner_product = (x_endclick - x1)*(x2 - x1) + (y_endclick - y1)*(y2 - y1)
            vector_norm = np.sqrt((x_endclick-x1)**2 + (y_endclick-y1)**2) * np.sqrt((x2-x1)**2 + (y2-y1)**2)
            cos_theta = inner_product/vector_norm
            width = (l * cos_theta) - np.sqrt((x2-x1)**2+(y2-y1)**2)
            x_scatter.append(width)

        We = 4.133 * np.std(x_scatter)

        for i in range(1, 15):
            time = [var for var in mat[3 * i + 3] if var is not None and isinstance(var, str) is False]
            time_init = time[0] / 1000
            time_end = time[len(time) - 1] / 1000
            x = [var for var in mat[3 * i + 1] if var is not None and isinstance(var, str) is False]
            y = [var for var in mat[3 * i + 2] if var is not None and isinstance(var, str) is False]
            x_init = x[0]
            y_init = y[0]
            x_end = x[len(x) - 1]
            y_end = y[len(y) - 1]

            dist = math.sqrt((x_end - x_init) ** 2 + (y_end - y_init) ** 2)
            tmp = np.log2(dist / We + 1) / (time_end - time_init)

            Throughput.append(tmp)
        return Throughput
    """

    def __CoRelation(self,TRE,TAC,MDC,ODC,MV,ME,MO,Throughput):

        tre = np.array(TRE)
        tac = np.array(TAC)
        mdc = np.array(MDC)
        odc = np.array(ODC)
        mv = np.array(MV)
        me = np.array(ME)
        mo = np.array(MO)
        tp = np.array(Throughput)

        datas = [tre, tac, mdc, odc, mv, me, mo, tp]
        datas = np.array(datas)

        rels = []
        for data in datas:
            tmp = []
            for element in datas:
                cor = np.corrcoef(element, data)[0, 1]
                tmp.append(cor)
            rels.append(tmp)

        return rels

    def __Saver_for_ModelComparison(self, TRE, TAC, MDC, ODC, MV, ME, MO, TP, path):
        self.ws['A1'].value = 'Attempt'
        self.ws['B1'].value = 'TRE'
        self.ws['C1'].value = 'TAC'
        self.ws['D1'].value = 'MDC'
        self.ws['E1'].value = 'ODC'
        self.ws['F1'].value = 'MV'

    def __Saver(self, TRE, TAC, MDC, ODC, MV, ME, MO, TP, path):
        if '1_1' in self.data_path:
            wb1 = px.Workbook()
        else:
            wb1 = px.load_workbook(path)
        ws1 = wb1.active

        ws1['A1'].value = 'Attempt'
        ws1['B1'].value = 'TRE'
        ws1['C1'].value = 'TAC'
        ws1['D1'].value = 'MDC'
        ws1['E1'].value = 'ODC'
        ws1['F1'].value = 'MV'
        ws1['G1'].value = 'ME'
        ws1['H1'].value = 'MO'
        ws1['I1'].value = 'Throughput'

        params = [TRE, TAC, MDC, ODC, MV, ME, MO, TP]
        alph = [chr(i) for i in range(65, 65+26)]

        """
        for i in range(len(params)):
            for j in range(len(params[i])):
                cell_num = alph[i+1] + str(j+2)
                ws1[cell_num].value = params[i][j]
        """
        # the below statement should be activated when file names are like 'attempt1_1'
        # then automatically this program examines TP

        for i in range(5):
            for j in range(5):
                order = str(i+1)+'_'+str(j+1)
                print(order)
                if order in self.data_path:
                    print('reached')
                    n = 5*i + (j+1)
                    for k in range(len(params)):
                        for l in range(len(params[k])):
                            cell_num = alph[k+1] + str(l+2 + (n-1)*14)
                            ws1['A' + str(l+2 + (n-1)*14)].value = n
                            ws1[cell_num].value = params[k][l]
                else:
                    pass

        # self.__analysis_whole(wb1, ws1)
        wb1.save(path)


        """
        for i in range(len(params)):
            for j in range(len(params[i])):
                cell_num = alph[i+1] + str(j+2+ (n-1)*14)
                ws1['A'+str(j+2 + (n-1)*14)].value = n
                ws1[cell_num].value = params[i][j]
        wb1.save(path)
        """

    def main(self):
        mat = self.__data_getter()
        TRE = self.__TRE_counter(mat)
        TAC = self.__Across_counter(mat)
        MDC = self.__Distance_from_line(mat)
        ODC = self.__Orthogonal_direction(mat)
        MV, ME, MO = self.__Movement_var_err_off(mat)
        Throughput = self.__Throughput_cal(mat)
        rels= self.__CoRelation(TRE, TAC, MDC, ODC, MV, ME, MO, Throughput)
        self.__Saver(TRE, TAC, MDC, ODC, MV, ME, MO, Throughput, self.path)

        return {'TRE': TRE, 'TAC': TAC, 'MDC': MDC, 'ODC': ODC, 'MV': MV, 'ME': ME, 'MO': MO, 'TP': Throughput}


if __name__ == '__main__':
    alph = [chr(i) for i in range(65, 65 + 26)] + ['A' + chr(i) for i in range(65, 65+26)]

    subjects = ['Inoue', 'Ishii', 'Kuramochi', 'Sudo', 'Tanaka']

    for subject in subjects:
        for i in range(5):
            for j in range(5):
                if subject == 'Tanaka':
                    if i == 1 and j == 2:
                        pass
                    else:
                        attempt = 'attempt' + str(i + 1) + '_' + str(j + 1) + '.xlsx'
                        data_path = 'C:/Users/socre/Google ドライブ/Pdev_results/experiment/results_data_2/' + subject + '/mouse/' + attempt
                        data_path_for_summary = 'C:/Users/socre/Google ドライブ/Pdev_results/experiment/results_data_2/' + subject + '/mouse/summed_2.xlsx'
                        test = Analyzer(data_path, data_path_for_summary)
                        test.main()
                else:
                    attempt = 'attempt' + str(i + 1) + '_' + str(j + 1) + '.xlsx'
                    data_path = 'C:/Users/socre/Google ドライブ/Pdev_results/experiment/results_data_2/' + subject + '/mouse/' + attempt
                    data_path_for_summary = 'C:/Users/socre/Google ドライブ/Pdev_results/experiment/results_data_2/' + subject + '/mouse/summed_2.xlsx'
                    test = Analyzer(data_path, data_path_for_summary)
                    test.main()

    """
    注意：田中のデータ2_3は欠損あり．
    """


    '''
    sudo_files = {'linear': '20181129_164537', 'tanh': '20181129_165023'}

    kitamura_files = {'linear': '20181129_185725', 'tanh': '20181129_190112'}

    tanaka_files = {'linear': '20190211_145725'}

    miya_files = {'linear': '20190211_150505'}


    data_path_for_summary = 'Miya_lin'
    path =''
    result_path = 'compare_lin_sig_tanh_dis/Tanaka_result.xlsx'

    sudo_params={}

    for key, item in tanaka_files.items():
        attempt = 'C:/Users/socre/Google ドライブ/Pdev_results/experiment/mobile_mover/results_data_4/TANAKA/' + key +\
                  '/WINDOW_' + item + '.xlsx'
        test = Analyzer(attempt, data_path_for_summary)
        params = test.main()
        sudo_params[key] = params
    try:
        wb = px.load_workbook(result_path)
        ws = wb.active
    except:
        wb = px.Workbook()
        ws = wb.active
    print(sudo_params)

    j = 0
    for key in sudo_params.keys():
        print(key)
        ws[alph[ws.max_column] + str(1)].value = key
        for mode in sudo_params[key].keys():
            ws[alph[j]+str(2)].value = mode
            i = 2
            for value in sudo_params[key][mode]:
                ws[alph[j]+str(i+1)].value = value
                i += 1
            j += 1

    wb.save(result_path)
    '''
    '''
    for key in sudo_params.keys():
        print('#######' + key + '############')
        ws[alph[ws.max_column] + str(1)].value = key
        ws[alph[ws.max_column] + str(1)].value = 'AVE'
        ws[alph[ws.max_column] + str(1)].value = 'STD'

        max_column = ws.max_column

        i = 2
        for param, item in sudo_params[key].items():
            ws[alph[max_column - 3] + str(i)].value = param
            ws[alph[max_column - 2] + str(i)].value = np.average(item)
            ws[alph[max_column - 1] + str(i)].value = np.std(item)
            i += 1

    # wb.save(result_path)
    '''
    """
    for i in range(5):
        # attempt = 'attempt1_' + str(i+1) + '.xlsx'
        data_path = 'C:/Users/socre/Google ドライブ/Pdev_results/experiment/results_data_3/Uchino_linear/' + attempt
        test = Analyzer(data_path, data_path_for_summary)
        test.main()
    """
