from pygame.locals import *
import pygame
import openpyxl as px
import numpy as np
import time
import math
import serial
from NN_calibulator import Layers_Net
import matplotlib.pyplot as plt
import datetime

class Drawer_continuous:

    def __init__(self, path='other/data_name.xlsx', fig_path='other/Figure1.png'):
        self.time_started = time.time()
        self.screen_size = (900, 800)
        pygame.init()
        self.screen = pygame.display.set_mode(self.screen_size)
        self.screen.fill((255, 255, 255))
        pygame.display.set_caption(u"angle configuration")

        self.path = path
        self.wb = px.Workbook()
        self.ws = self.wb.active
        self.time = [0]
        self.fig_path = fig_path

        self.center = (450, 750)
        self.duration = 800
        self.Theta_in = []
        self.Theta_label = []
        self.Theta_ = []

        self.ser = serial.Serial('COM10', 19200)
        self.calibrator = Layers_Net()
        self.start = time.time()
        self.int_start = int(time.time())
        self.flag = 0
        self.occur_time = None

    def __theta(self, i):
        i = i % (2 * self.duration)
        if i <= self.duration:
            return np.pi * i / self.duration
        else:
            return np.pi * (2 - i / self.duration)

    def __getter_angle(self):
        datas = self.ser.readline()
        datas = datas.decode().split(';')
        try:
            ax = int(datas[0]) / 16384
            ay = int(datas[1]) / 16384
            az = int(datas[2]) / 16384
        except:
            ax, ay, az = 0, 0, 0

        acc_pitch = math.atan2(ax, math.sqrt(az ** 2 + ay ** 2)) * 360 / 2 / math.pi
        acc_roll = math.atan2(ay, math.sqrt(ax ** 2 + az ** 2)) * 360 / 2 / math.pi

        return -1*acc_pitch, -1*acc_roll

    def __center_point(self):
        pygame.draw.circle(self.screen, (0, 0, 0), self.center, 10, 2)
        pygame.draw.line(self.screen, (0, 0, 0), (0, self.center[1]+10), (900, self.center[1]+10), 5)

    def __line_drawer(self, theta):
        x_nxt = 450 - 400 * np.cos(theta)
        y_nxt = 750 - 400 * abs(np.sin(theta))
        pygame.draw.line(self.screen, (20, 138, 20), self.center, (x_nxt, y_nxt), 5)

    def __line_remover(self, theta):
        x_prev = 450 - 400 * np.cos(theta)
        y_prev = 750 - 400 * abs(np.sin(theta))
        pygame.draw.line(self.screen, (255, 255, 255), self.center, (x_prev, y_prev), 5)

    def __saver(self, i, data_path='other/data_name.xlsx'):
        self.ws['I1'].value = 'count'
        self.ws['J1'].value = 'Theta_input'
        self.ws['K1'].value = 'Theta_label'
        self.ws['L1'].value = 'Theta_'
        self.ws['M1'].value = 'Time'

        for j in range(i):
            i_cell = 'I' + str(j+2)
            Theta_in_cell = 'J' + str(j+2)
            Theta_label_cell = 'K' + str(j+2)
            Theta_cell = 'L' + str(j+2)
            Time_cell = 'M' + str(j+2)

            self.ws[i_cell].value = j
            self.ws[Theta_in_cell].value = self.Theta_in[j]
            self.ws[Theta_label_cell].value = self.Theta_label[j]
            self.ws[Theta_cell].value = self.Theta_[j]
            self.ws[Time_cell].value = self.time[j]

        self.wb.save(data_path)

    def __plot(self, x_datas,  Theta_in, Theta_label, Theta_):
        del x_datas[-1]
        fig = plt.figure()
        plt.plot(x_datas, Theta_in, lw=2, label='input angle')
        # plt.plot(x_datas, Theta_label, lw=2, label='ideal angle')
        # plt.plot(x_datas, Theta_, lw=2, label='output from NN')
        plt.tick_params(labelsize=14)

        plt.legend(bbox_to_anchor=(1, 1), loc='upper right', fontsize=18)
        plt.ylim(-180, 180)
        plt.title('Angle-Time', fontsize=28)
        plt.xlabel('time(sec)', fontsize=20)
        plt.ylabel('angle(deg)', fontsize=20)
        plt.grid(True)
        plt.savefig(self.fig_path, bbox_inches='tight')
        plt.show()

    def __time_display(self, i):
        sysfont = pygame.font.SysFont(None, 80)

        if i == 0:
            pass
        else:
            cnt_del = sysfont.render('time passed: ' + str(int(self.time[i-1])), False, (255, 255, 255))
            self.screen.blit(cnt_del, (20, 20))

        cnt = sysfont.render('time passed: '+str(int(self.time[i])), False, (0, 0, 0))
        self.screen.blit(cnt, (20, 20))

    def main(self):
        self.__center_point()
        i = 0
        while True:
            now = time.time() - self.start
            label_x = (self.__theta(i) * 180 / np.pi)
            if i > 1:
                self.__line_remover(self.__theta(i-1))
            self.__line_drawer(self.__theta(i))
            theta_x, theta_y = self.__getter_angle()
            print(theta_x, theta_y)
            theta_in = np.array([theta_x, theta_y]).reshape(1, 2)
            theta_label = np.array([label_x - 90, 0]).reshape(1, 2)
            theta_ = self.calibrator.predict(theta_in)
            residue = self.calibrator.train_params(theta_, theta_label)
            self.Theta_in.append(theta_in[0][1])
            # self.Theta_in.append(theta_in[0][0])
            self.Theta_label.append(theta_label[0][0])
            self.Theta_.append(theta_[0][0])
            self.time.append(now)
            self.__time_display(i)
            for event in pygame.event.get():
                if event.type == QUIT:
                    self.__saver(i, self.path)
                    self.calibrator.save_params()
                    self.__plot(self.time, self.Theta_in, self.Theta_label, self.Theta_)
                    quit()

            pygame.display.update()
            i += 1

if __name__ == '__main__':
    now = 'other/PERSON/' + str(datetime.datetime.now().strftime('%Y%m%d_%H%M%S')) + '.xlsx'
    fig_path = 'other/PERSON/' + str(datetime.datetime.now().strftime('%Y%m%d_%H%M%S')) + '.png'
    test = Drawer_continuous(now, fig_path)
    test.main()
