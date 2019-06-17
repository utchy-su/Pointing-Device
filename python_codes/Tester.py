# -*- coding: utf-8 -*-
"""
Created on Mon Jun 11 13:58:11 2018

@author: smart
"""

# coding:utf-8
import pygame
from pygame.locals import *
import numpy as np
import math
import openpyxl as px
import keyboard
from Serial_Communicate import MySerial
import time as timer


class Drawer:

    def __init__(self, data_path, data_path_mover):
        self.screen_size = (900, 900)
        pygame.init()
        self.screen = pygame.display.set_mode(self.screen_size)
        self.screen.fill((255, 255, 255))
        pygame.display.set_caption(u"mouse configuration")
        #########
        self.wb = px.Workbook()
        self.ws = self.wb.active
        self.x = []
        self.y = []
        self.time = []
        self.diameter = 30
        self.data_path = data_path
        self.data_path_mover = data_path_mover

        self.serialRead = MySerial('COM13')
        self.angle_X = []
        self.angle_Y = []

    def __drawer_circle(self):
        pygame.draw.circle(self.screen, (0, 0, 0), (450, 450), 200, 2)
        for i in range(16):
            x = 450 + 200 * np.cos(math.pi*(i/8))
            y = 450 + 200 * np.sin(math.pi*(i/8))
            pygame.draw.circle(self.screen, (255, 255, 255), (int(x), int(y)), self.diameter)
            pygame.draw.circle(self.screen, (0, 0, 0), (int(x), int(y)), self.diameter, 2)

    def __responder(self, click_count, order_now, order):
        x_now, y_now = pygame.mouse.get_pos()
        if math.sqrt((x_now-450-200*np.cos(math.pi*order_now/8))**2+(y_now-450-200*np.sin(math.pi*order_now/8))**2)<=self.diameter:
            if self.__area_distinguisher() <= 5:
                ##############
                if click_count >= 1:
                    self.__tgt_circle_remover(order[click_count-1])
                    self.__tgt_line_remover(click_count, order[click_count], order)
                if click_count <= 13:
                    self.__tgt_circle(order[click_count+1])
                    self.__tgt_line(click_count, order_now, order)
                else:
                    pass
                ##############

                click_count += 1
                detect = 1
            else:
                detect = 0
        else:
            detect = 0

        return click_count, detect

    def __area_distinguisher(self):

        '''speed distinguisher
        not perfectly implemented yet
        the way using time that a poninter stay in the cicle is affecting the value of TP
        '''

        try:
            x_vel = (float(self.x[-1]) - float(self.x[-2]))/(float(self.time[-1])-float(self.time[-2]))*1000
            y_vel = (float(self.y[-1]) - float(self.y[-2]))/(float(self.time[-1])-float(self.time[-2]))*1000

            vel = math.sqrt(x_vel**2 + y_vel**2)
        except:
            vel = 0
        return vel

    def __tgt_circle(self, order_now):
        x = 450 + 200*np.cos(math.pi*(order_now/8))
        y = 450 + 200*np.sin(math.pi*(order_now/8))
        pygame.draw.circle(self.screen, (255, 0, 0), (int(x), int(y)), self.diameter)

    def __tgt_circle_remover(self, order_now):
        x = 450 + 200*np.cos(math.pi*(order_now/8))
        y = 450 + 200*np.sin(math.pi*(order_now/8))
        pygame.draw.circle(self.screen, (255, 255, 255), (int(x), int(y)), self.diameter)
        pygame.draw.circle(self.screen, (0, 0, 0), (int(x), int(y)), self.diameter, 2)

    def __tgt_line(self, click_count, order_now, order):
        x = 450 + 200*np.cos(math.pi*(order_now/8))
        y = 450 + 200*np.sin(math.pi*(order_now/8))
        x_nxt = 450 + 200*np.cos(math.pi*(order[click_count+1])/8)
        y_nxt = 450 + 200*np.sin(math.pi*(order[click_count+1])/8)

        pygame.draw.line(self.screen, (20, 128, 20), (x, y), (x_nxt, y_nxt), 5)

    def __tgt_line_remover(self, click_count, order_now, order):
        x = 450 + 200*np.cos(math.pi*(order_now/8))
        y = 450 + 200*np.sin(math.pi*(order_now/8))
        x_prev = 450 + 200*np.cos(math.pi*(order[click_count-1])/8)
        y_prev = 450 + 200*np.sin(math.pi*(order[click_count-1])/8)

        pygame.draw.line(self.screen,(255,255,255),(x_prev,y_prev),(x,y),5)

        pygame.draw.circle(self.screen,(0,0,0),(int(x_prev),int(y_prev)),self.diameter,2)

    def __recorder(self, detect):
        now = pygame.time.get_ticks()
        x_now, y_now = pygame.mouse.get_pos()
        angX, angY, _ = self.serialRead.get_angle()

        if detect == 1:
            self.x.append('_')
            self.y.append('_')
            self.angle_X.append('_')
            self.angle_Y.append('_')
            self.time.append('_')
            self.x.append(str(x_now))
            self.y.append(str(y_now))
            self.time.append(str(now))
            self.angle_X.append(str(angX))
            self.angle_Y.append(str(angY))
        else:
            self.x.append(str(x_now))
            self.y.append(str(y_now))
            self.time.append(str(now))
            self.angle_X.append(str(angX))
            self.angle_Y.append(str(angY))

    def __saver(self, order):
        time_ser = ','.join(self.time)
        x_ser = ','.join(self.x)
        y_ser = ','.join(self.y)
        ang_X_ser = ','.join(self.angle_X)
        ang_Y_ser = ','.join(self.angle_Y)

        time_listed = time_ser.split('_')
        x_listed = x_ser.split('_')
        y_listed = y_ser.split('_')
        ang_X_listed = ang_X_ser.split('_')
        ang_Y_listed = ang_Y_ser.split('_')

        alph = [chr(i) for i in range(65, 65+26)]
        alph_to_append = ['A'+chr(i) for i in range(65, 65+26)]
        alph_to_append2 = ['B'+chr(i) for i in range(65, 65+26)]

        alph += alph_to_append + alph_to_append2

        for i in range(15):
            time = time_listed[i].split(',')
            xroute = x_listed[i].split(',')
            yroute = y_listed[i].split(',')
            ang_x_ = ang_X_listed[i].split(',')
            ang_y_ = ang_Y_listed[i].split(',')

            del time[0]
            del xroute[0]
            del yroute[0]
            del ang_x_[0]
            del ang_y_[0]

            for j in range(len(order)):
                order_cell = 'A' + str(j+2)
                self.ws[order_cell].value = order[j]
                self.ws['A1'].value = 'orders'

            for j in range(len(time)-1):
                namex= alph[5*i+1]+str(1)
                namey = alph[5*i+2]+str(1)
                name_angx = alph[5*i+3]+str(1)
                name_angy = alph[5*i+4]+str(1)
                namet = alph[5*i+5] + str(1)

                x_cell = alph[5*i+1]+str(j+2)
                y_cell = alph[5*i+2]+str(j+2)
                ang_x_cell = alph[5*i+3] + str(j+2)
                ang_y_cell = alph[5*i+4] + str(j+2)
                time_cell = alph[5*i+5]+str(j+2)

                self.ws[namex] = 'x from ' + str(i) + ' to ' + str(i+1)
                self.ws[namey] = 'y from ' + str(i) + ' to ' + str(i+1)
                self.ws[name_angx] = 'angle X during' + str(i) + 'to' + str(i+1)
                self.ws[name_angy] = 'angle Y during' + str(i) + 'to' + str(i+1)
                self.ws[namet] = 'time from ' + str(i) + ' to ' + str(i+1)

                self.ws[x_cell].value = xroute[j]
                self.ws[y_cell].value = yroute[j]
                self.ws[ang_x_cell].value = ang_x_[j]
                self.ws[ang_y_cell].value = ang_y_[j]
                self.ws[time_cell].value = time[j]


        self.wb.save(self.data_path)

    def __count_display(self, click_count):
        sysfont = pygame.font.SysFont(None,80)
        if click_count == 0:
            pass
        else:
            cnt_del = sysfont.render('click count='+str(click_count-1), False, (255, 255, 255))
            self.screen.blit(cnt_del, (20, 20))

        cnt = sysfont.render('click count='+str(click_count), False, (0, 0, 0))
        self.screen.blit(cnt, (20, 20))

        if click_count == 15:
            fin = sysfont.render('Finished', False, (0, 0, 0))
            self.screen.blit(fin, (80, 80))


    def main(self):
        a = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
        i = 1
        click_count=0
        # order = np.random.choice(a,15,replace=False)
        # order = [4,0,15,10,5]
        order = [0, 8, 1, 9, 2, 10, 3, 11, 4, 12, 5, 13, 6, 14, 7, 15]
        self.__drawer_circle()
        self.__tgt_circle(order[click_count])
        try:
            _, _, gain = self.serialRead.get_angle()
            print("gain: ", gain)
        except FileNotFoundError:
            print("mode: None")
        while True:
            detect = 0
            self.__count_display(click_count)
            for event in pygame.event.get():
                if event.type == QUIT:
                    quit()

                if event.type == KEYDOWN or (event.type == MOUSEBUTTONDOWN and event.button == 1):
                    click_count, detect = self.__responder(click_count, order[click_count], order)
                    #print(click_count)
                else:
                    pass
            self.__recorder(detect)

            if click_count == 15:
                self.__count_display(click_count)
                pygame.display.update()
                self.__saver(order)
                self.serialRead.close()
                break

            pygame.display.update()
            i += 1


if __name__ == '__main__':
    for attempt in range(5):
        name = 'attempt' + str(attempt+1)
        data_path = 'sensitivity_tuning/nonlin_90/' + name + '.xlsx'
        test = Drawer(data_path, 'N/A')
        test.main()
        print("if ready for the next, press ESC key")
        while True:
            if keyboard.is_pressed("esc") == 1:
                break
        del test

    """
    for attempt in range(25):
        name = 'attempt' + str(attempt+1)
        data_path = 'C:/Users/socre/Google ドライブ/Pdev_results/experiment/linear vs non-linear result/Sasaki/linear/' + name + '.xlsx'
        test = Drawer(data_path, 'N/A')
        test.main()
        print('Attempt finished\n')
        print('If ready for the next, press ESC key')
        while True:
            if keyboard.is_pressed('esc') == 1:
                break
        del test
    """